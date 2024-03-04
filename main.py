import openpyxl as xl
from openpyxl.styles import Font

inventory_file = xl.load_workbook("inventory.xlsx")
sheet = inventory_file["Sheet1"]

products_per_supplier = {}
products_under_than_10_inventory = {}
total_inventory_value_per_supplier = {}

sheet.cell(1, 5).value = "Inventory Value"
sheet.cell(1, 5).font = Font(bold=True)

for product_row in range(2, sheet.max_row + 1):
    product = int(sheet.cell(product_row, 1).value)
    inventory = int(sheet.cell(product_row, 2).value)
    price = sheet.cell(product_row, 3).value
    supplier = sheet.cell(product_row, 4).value
    inventory_value = sheet.cell(product_row, 5)
    
    # Calculate and list number of products per supplier/company
    if supplier in products_per_supplier:
        products_per_supplier[supplier] = products_per_supplier[supplier] + 1
    else:
        print(f"Adding supplier {supplier}...")
        products_per_supplier[supplier] = 1
        
    # List products with inventory less than 10
    if inventory < 10:
        products_under_than_10_inventory[product] = inventory
        
    # Calculate and list suppliers with total inventory values
    if supplier in total_inventory_value_per_supplier:
        total_inventory_value_per_supplier[supplier] = total_inventory_value_per_supplier[supplier] + (inventory * price)
    else:
        total_inventory_value_per_supplier[supplier] = inventory * price
        
    # Write inventory value to each product in an additional column
    inventory_value.value = inventory * price
    
inventory_file.save("inventory_with_total_value.xlsx")

print(f"\nProducts per supplier:\n{products_per_supplier}")
print(f"\nProducts with less than 10 inventory value:\n{products_under_than_10_inventory}")
print(f"\nTotal inventory value per supplier:\n{total_inventory_value_per_supplier}")